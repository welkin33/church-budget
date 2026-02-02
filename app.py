"""
Church Worship Team Budget Manager
Zero-cost serverless web app using Streamlit + Google Sheets + Google Drive + Vision OCR
"""

import io
import re
import json
import datetime
from typing import Optional

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ──────────────────────────────────────────────
# Page Config
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="고등부 예산 관리",
    page_icon="⛪",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# Custom CSS — Sophisticated dark theme
# ──────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

:root {
    --primary: #6C63FF;
    --primary-light: #8B83FF;
    --accent: #00D2FF;
    --success: #00E676;
    --warning: #FFD600;
    --danger: #FF5252;
    --bg-dark: #0E1117;
    --bg-card: #1A1D26;
    --bg-card-hover: #22262F;
    --text-primary: #FAFAFA;
    --text-secondary: #A0AEC0;
    --border: #2D3748;
}

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Header gradient */
.main-header {
    background: linear-gradient(135deg, #6C63FF 0%, #00D2FF 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 2.4rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
    letter-spacing: -0.02em;
}

.sub-header {
    color: var(--text-secondary);
    font-size: 1rem;
    font-weight: 300;
    margin-bottom: 2rem;
}

/* Metric cards */
.metric-card {
    background: linear-gradient(145deg, #1A1D26 0%, #22262F 100%);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 1.5rem;
    text-align: center;
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}
.metric-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(108, 99, 255, 0.15);
}
.metric-label {
    color: var(--text-secondary);
    font-size: 0.85rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 0.5rem;
}
.metric-value {
    font-size: 1.8rem;
    font-weight: 700;
    color: var(--text-primary);
}
.metric-value.primary { color: var(--primary-light); }
.metric-value.success { color: var(--success); }
.metric-value.warning { color: var(--warning); }
.metric-value.danger  { color: var(--danger); }

/* Upload area */
.upload-zone {
    border: 2px dashed var(--primary);
    border-radius: 16px;
    padding: 2rem;
    text-align: center;
    background: rgba(108, 99, 255, 0.05);
    margin: 1rem 0;
}

/* Section divider */
.section-title {
    font-size: 1.3rem;
    font-weight: 600;
    background: linear-gradient(135deg, #6C63FF 0%, #00D2FF 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin: 2rem 0 1rem 0;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid var(--primary);
    display: inline-block;
}

/* Sidebar styling */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0E1117 0%, #1A1D26 100%);
}

/* Table styling */
.dataframe {
    border-radius: 12px !important;
    overflow: hidden;
}

/* Button styling */
.stButton > button {
    background: linear-gradient(135deg, #6C63FF 0%, #8B83FF 100%);
    color: white;
    border: none;
    border-radius: 10px;
    padding: 0.6rem 1.5rem;
    font-weight: 600;
    transition: all 0.2s ease;
}
.stButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 15px rgba(108, 99, 255, 0.4);
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px;
    padding: 8px 20px;
    font-weight: 500;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# Google API Connections (cached)
# ──────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/cloud-platform",
]


@st.cache_resource(ttl=300)
def get_credentials():
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return creds


@st.cache_resource(ttl=300)
def get_gspread_client():
    return gspread.authorize(get_credentials())


@st.cache_resource(ttl=300)
def get_drive_service():
    return build("drive", "v3", credentials=get_credentials())


@st.cache_resource(ttl=300)
def get_vision_client():
    from google.cloud import vision
    creds = get_credentials()
    return vision.ImageAnnotatorClient(credentials=creds)


def get_sheet():
    client = get_gspread_client()
    sheet_id = st.secrets["app_settings"]["spreadsheet_id"]
    spreadsheet = client.open_by_key(sheet_id)
    try:
        ws = spreadsheet.worksheet("Transactions")
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="Transactions", rows=1000, cols=10)
        ws.append_row([
            "Date", "Category", "Description", "Amount",
            "Payment Method", "Receipt URL", "OCR Amount", "Submitted By", "Timestamp"
        ])
    return ws


def get_budget_sheet():
    client = get_gspread_client()
    sheet_id = st.secrets["app_settings"]["spreadsheet_id"]
    spreadsheet = client.open_by_key(sheet_id)
    try:
        ws = spreadsheet.worksheet("Budget")
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="Budget", rows=100, cols=5)
        ws.append_row(["Category", "Monthly Budget", "Year", "Notes"])
    return ws


# ──────────────────────────────────────────────
# OCR via Google Cloud Vision
# ──────────────────────────────────────────────
def extract_amount_from_image(image_bytes: bytes) -> tuple[Optional[str], Optional[str]]:
    """Use Google Vision OCR to extract total amount from Korean receipt.
    Returns (amount_string, full_ocr_text) for debugging.
    """
    from google.cloud import vision

    client = get_vision_client()
    image = vision.Image(content=image_bytes)

    # document_text_detection is much better for structured docs like receipts
    response = client.document_text_detection(
        image=image,
        image_context={"language_hints": ["ko", "en"]},
    )

    if response.error.message:
        return None, None

    full_text = response.full_text_annotation.text if response.full_text_annotation else None
    if not full_text:
        return None, None

    lines = full_text.split("\n")
    keyword_candidates = []

    # Pattern: Korean won amounts like "27,190원" or "27,190" (with commas, reasonable length)
    # Max 10 digits to avoid picking up order numbers, phone numbers, etc.
    won_pattern = re.compile(r'([\d,]{1,12})(?:원)?')

    # Korean receipt keywords — must end with "금액" or be exact total keywords
    # to avoid matching "승인번호", "결제일시" etc.
    total_keywords = [
        "합계금액", "합계 금액", "총합계", "총 합계",
        "결제금액", "결제 금액", "승인금액", "승인 금액",
        "합계", "합 계",
        "총액", "총 액", "합산",
        "받을금액", "받을 금액", "청구금액", "청구 금액",
        "total", "amount",
    ]

    # Keywords that should NOT match (to filter "승인번호", "거래일시" etc.)
    exclude_keywords = ["번호", "일시", "종류", "개월", "상호", "주소", "등록", "상점", "정보"]

    def extract_won_amount(text: str) -> list[int]:
        """Extract reasonable Korean won amounts from text (100 ~ 99,999,999)."""
        results = []
        for m in won_pattern.finditer(text):
            raw = m.group(1).replace(",", "")
            if raw.isdigit():
                val = int(raw)
                if 100 <= val <= 99_999_999:
                    results.append(val)
        return results

    def line_has_exclude(text: str) -> bool:
        for ex in exclude_keywords:
            if ex in text:
                return True
        return False

    # Strategy 1: Find "합계금액" and look for amount with "원" suffix nearby
    # Handle receipts where labels and values are in separate blocks
    won_with_suffix = re.compile(r'([\d,]{1,12})원')
    for i, line in enumerate(lines):
        stripped = line.strip().replace(" ", "")
        if "합계금액" in stripped or "합계 금액" in stripped.replace(" ", ""):
            # Look for "XX,XXX원" pattern in same line
            m = won_with_suffix.search(line)
            if m:
                val = int(m.group(1).replace(",", ""))
                if 100 <= val <= 99_999_999:
                    return f"{val:,.0f}", full_text
            # Not on same line — scan ALL following lines for the first "원" suffixed amount
            for j in range(i + 1, len(lines)):
                m = won_with_suffix.search(lines[j])
                if m:
                    val = int(m.group(1).replace(",", ""))
                    if 100 <= val <= 99_999_999:
                        return f"{val:,.0f}", full_text

    # Strategy 2: General keyword search (same line or next line)
    for i, line in enumerate(lines):
        stripped = line.strip().replace(" ", "")
        if line_has_exclude(stripped):
            continue
        for priority, kw in enumerate(total_keywords):
            kw_clean = kw.replace(" ", "")
            if kw_clean in stripped:
                # Same line first
                amounts = extract_won_amount(line)
                if amounts:
                    for val in amounts:
                        keyword_candidates.append((priority, val))
                else:
                    # Next line
                    if i + 1 < len(lines) and not line_has_exclude(lines[i + 1]):
                        amounts = extract_won_amount(lines[i + 1])
                        for val in amounts:
                            keyword_candidates.append((priority, val))

    if keyword_candidates:
        keyword_candidates.sort(key=lambda x: (x[0], -x[1]))
        best_priority = keyword_candidates[0][0]
        best_group = [v for p, v in keyword_candidates if p == best_priority]
        return f"{max(best_group):,.0f}", full_text

    # Fallback: largest reasonable amount in entire text
    fallback = extract_won_amount(full_text)

    if fallback:
        return f"{max(fallback):,.0f}", full_text
    return None, full_text


# ──────────────────────────────────────────────
# Google Drive Upload
# ──────────────────────────────────────────────
def upload_to_drive(file_bytes: bytes, filename: str, mime_type: str) -> str:
    """Upload file to Google Drive shared folder, return public URL."""
    service = get_drive_service()
    folder_id = st.secrets["app_settings"]["drive_folder_id"]

    file_metadata = {
        "name": filename,
        "parents": [folder_id],
    }
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=True)

    # supportsAllDrives allows upload into folders shared with the service account
    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True,
    ).execute()

    # Make publicly viewable
    try:
        service.permissions().create(
            fileId=file["id"],
            body={"type": "anyone", "role": "reader"},
            supportsAllDrives=True,
        ).execute()
    except Exception:
        pass  # folder-level sharing may already cover this

    return file.get("webViewLink", f"https://drive.google.com/file/d/{file['id']}/view")


# ──────────────────────────────────────────────
# Data Loading
# ──────────────────────────────────────────────
@st.cache_data(ttl=60)
def load_transactions() -> pd.DataFrame:
    ws = get_sheet()
    data = ws.get_all_records()
    if not data:
        return pd.DataFrame(columns=[
            "Date", "Category", "Description", "Amount",
            "Payment Method", "Receipt URL", "OCR Amount", "Submitted By", "Timestamp"
        ])
    df = pd.DataFrame(data)
    if "Amount" in df.columns:
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df


@st.cache_data(ttl=60)
def load_budgets() -> pd.DataFrame:
    ws = get_budget_sheet()
    data = ws.get_all_records()
    if not data:
        return pd.DataFrame(columns=["Category", "Monthly Budget", "Year", "Notes"])
    df = pd.DataFrame(data)
    df["Monthly Budget"] = pd.to_numeric(df["Monthly Budget"], errors="coerce").fillna(0)
    return df


# ──────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────
def generate_excel_report(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df = df.copy()
        if "Date" in export_df.columns:
            export_df["Date"] = export_df["Date"].astype(str)
        export_df.to_excel(writer, index=False, sheet_name="Transactions")

        # Add hyperlinks for receipt URLs
        ws = writer.sheets["Transactions"]
        url_col = None
        for idx, col in enumerate(export_df.columns, 1):
            if col == "Receipt URL":
                url_col = idx
                break
        if url_col:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(url_col)
            for row_idx in range(2, len(export_df) + 2):
                cell = ws[f"{col_letter}{row_idx}"]
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        # Summary sheet
        if not df.empty and "Category" in df.columns and "Amount" in df.columns:
            summary = df.groupby("Category")["Amount"].agg(["sum", "count", "mean"]).reset_index()
            summary.columns = ["Category", "Total", "Count", "Average"]
            summary.to_excel(writer, index=False, sheet_name="Summary")

    return output.getvalue()


# ──────────────────────────────────────────────
# UI Helper
# ──────────────────────────────────────────────
def metric_card(label: str, value: str, color_class: str = "primary"):
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value {color_class}">{value}</div>
    </div>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p class="main-header" style="font-size:1.6rem;">⛪ 고등부 예산</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header" style="font-size:0.85rem;">Church Worship Team Budget</p>', unsafe_allow_html=True)
    st.divider()

    page = st.radio(
        "메뉴 / Menu",
        ["📊 대시보드", "📤 지출 입력", "📋 거래 내역", "⚙️ 예산 설정", "📥 리포트 다운로드"],
        label_visibility="collapsed",
    )
    st.divider()
    if st.button("🔄 데이터 새로고침"):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")
    st.caption("Powered by Streamlit · Google Sheets · Vision AI")

# ──────────────────────────────────────────────
# PAGE: Dashboard
# ──────────────────────────────────────────────
if page == "📊 대시보드":
    st.markdown('<p class="main-header">Budget Dashboard</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">고등부 예산 현황을 한눈에 확인하세요</p>', unsafe_allow_html=True)

    df = load_transactions()
    budgets = load_budgets()

    # Calculate metrics regardless of df being empty
    now = datetime.datetime.now()
    current_month = now.month
    current_year = now.year

    monthly_budget = budgets["Monthly Budget"].sum() if not budgets.empty else 0

    # 총예산 = 이번달 예산 + 이월 잔액 (unspent budget from previous months this year)
    months_elapsed = current_month  # including current month
    cumulative_budget = monthly_budget * months_elapsed

    if not df.empty and "Date" in df.columns:
        # This month's spending
        this_month_mask = (df["Date"].dt.month == current_month) & (df["Date"].dt.year == current_year)
        this_month_spent = df.loc[this_month_mask, "Amount"].sum()

        # Total spending (all time)
        total_spent = df["Amount"].sum()

        # Previous months spent this year (for carryover calc)
        this_year_mask = df["Date"].dt.year == current_year
        total_spent_this_year = df.loc[this_year_mask, "Amount"].sum()

        tx_count = len(df)
    else:
        this_month_spent = 0
        total_spent = 0
        total_spent_this_year = 0
        tx_count = 0

    # 총예산 = 올해 누적 예산 (이번달 예산 + 이월된 예산)
    total_budget = cumulative_budget
    remaining = total_budget - total_spent_this_year

    # KPI Cards — 6 metrics
    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card("이번달 예산", f"₩{monthly_budget:,.0f}", "primary")
    with c2:
        metric_card("총예산 (이번달+이월)", f"₩{total_budget:,.0f}", "primary")
    with c3:
        metric_card("이번달 지출", f"₩{this_month_spent:,.0f}", "warning")

    c4, c5, c6 = st.columns(3)
    with c4:
        metric_card("총 지출", f"₩{total_spent:,.0f}", "danger")
    with c5:
        metric_card("잔여 예산", f"₩{remaining:,.0f}", "success" if remaining > 0 else "danger")
    with c6:
        metric_card("거래 건수", f"{tx_count}건", "warning")

    if df.empty:
        st.info("아직 등록된 거래가 없습니다. '지출 입력' 메뉴에서 첫 거래를 등록하세요!")
    else:

        st.markdown("")

        # Charts row
        col_left, col_right = st.columns(2)

        with col_left:
            st.markdown('<p class="section-title">카테고리별 지출</p>', unsafe_allow_html=True)
            if "Category" in df.columns:
                cat_data = df.groupby("Category")["Amount"].sum().reset_index()
                fig = px.pie(
                    cat_data, values="Amount", names="Category",
                    color_discrete_sequence=["#6C63FF", "#FF5252", "#00D2FF", "#00E676", "#FFD600", "#FF6E40", "#AB47BC", "#26C6DA"],
                    hole=0.45,
                )
                fig.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#FAFAFA"),
                    legend=dict(font=dict(size=12)),
                    margin=dict(t=20, b=20, l=20, r=20),
                )
                st.plotly_chart(fig, config={"displayModeBar": False})

        with col_right:
            st.markdown('<p class="section-title">예산 대비 지출 현황</p>', unsafe_allow_html=True)
            if not budgets.empty:
                cat_spent = df.groupby("Category")["Amount"].sum().reset_index()
                comparison = budgets.merge(cat_spent, on="Category", how="left").fillna(0)
                comparison["Budget"] = comparison["Monthly Budget"] * months_elapsed
                comparison["Spent"] = comparison["Amount"]
                comparison["Remaining"] = (comparison["Budget"] - comparison["Spent"]).clip(lower=0)
                comparison["Over"] = (comparison["Spent"] - comparison["Budget"]).clip(lower=0)

                fig2 = go.Figure()
                fig2.add_trace(go.Bar(
                    name="지출",
                    x=comparison["Category"],
                    y=comparison["Spent"] - comparison["Over"],
                    marker_color="#FF5252",
                    text=comparison["Spent"].apply(lambda x: f"₩{x:,.0f}"),
                    textposition="inside",
                    textfont=dict(color="#FAFAFA"),
                ))
                fig2.add_trace(go.Bar(
                    name="초과",
                    x=comparison["Category"],
                    y=comparison["Over"],
                    marker_color="#FF1744",
                    textfont=dict(color="#FAFAFA"),
                ))
                fig2.add_trace(go.Bar(
                    name="잔여",
                    x=comparison["Category"],
                    y=comparison["Remaining"],
                    marker_color="#8B83FF",
                    text=comparison["Remaining"].apply(lambda x: f"₩{x:,.0f}"),
                    textposition="inside",
                    textfont=dict(color="#FAFAFA"),
                ))
                fig2.update_layout(
                    barmode="stack",
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#FAFAFA"),
                    xaxis=dict(gridcolor="#2D3748"),
                    yaxis=dict(gridcolor="#2D3748", title="원"),
                    margin=dict(t=30, b=20, l=20, r=20),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                )
                st.plotly_chart(fig2, config={"displayModeBar": False})


# ──────────────────────────────────────────────
# PAGE: Expense Input (with Receipt Upload + OCR)
# ──────────────────────────────────────────────
elif page == "📤 지출 입력":
    st.markdown('<p class="main-header">지출 입력</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">영수증을 업로드하면 금액이 자동 인식됩니다</p>', unsafe_allow_html=True)

    budgets = load_budgets()
    categories = budgets["Category"].tolist() if not budgets.empty else [
        "악기/장비", "음향장비", "악보/교재", "식비/간식", "교통비", "기타"
    ]

    # Receipt Upload Section
    st.markdown('<p class="section-title">📸 영수증 업로드 (선택)</p>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        "영수증 이미지를 드래그하거나 클릭하여 업로드",
        type=["jpg", "jpeg", "png", "webp"],
        help="JPG, PNG, WEBP 형식 지원",
    )

    ocr_amount = ""
    receipt_url = ""

    if uploaded_file is not None:
        col_img, col_ocr = st.columns([1, 1])

        with col_img:
            st.image(uploaded_file, caption="업로드된 영수증", width="stretch")

        with col_ocr:
            file_bytes = uploaded_file.getvalue()

            with st.spinner("🔍 OCR 분석 중..."):
                try:
                    detected, ocr_text = extract_amount_from_image(file_bytes)
                    if detected:
                        ocr_amount = detected
                        st.success(f"✅ 인식된 금액: **₩{detected}**")
                        st.caption("금액이 정확하지 않으면 아래에서 직접 수정하세요.")
                    else:
                        st.warning("금액을 자동 인식하지 못했습니다. 직접 입력해주세요.")
                    if ocr_text:
                        with st.expander("🔎 OCR 인식 원문 보기"):
                            st.text(ocr_text)
                except Exception as e:
                    st.warning(f"OCR 처리 중 오류: {e}")
                    st.caption("금액을 직접 입력해주세요.")

    # Input Form
    st.markdown('<p class="section-title">📝 거래 정보</p>', unsafe_allow_html=True)

    with st.form("expense_form", clear_on_submit=True):
        col1, col2 = st.columns(2)

        with col1:
            date = st.date_input("날짜", value=datetime.date.today())
            category = st.selectbox("카테고리", categories)
            description = st.text_input("설명", placeholder="예: 건반 수리비")

        with col2:
            default_amount = int(float(ocr_amount.replace(",", ""))) if ocr_amount else 0
            amount = st.number_input(
                "금액 (원)",
                min_value=0,
                value=default_amount,
                step=1000,
                help="OCR로 인식된 금액이 자동 입력됩니다",
            )
            payment_method = st.selectbox("결제 수단", ["카드", "현금", "계좌이체", "기타"])
            submitted_by = st.text_input("입력자", placeholder="이름")

        submitted = st.form_submit_button("💾 저장하기", width="stretch")

        if submitted:
            if amount <= 0:
                st.error("금액을 입력해주세요.")
            elif not description:
                st.error("설명을 입력해주세요.")
            else:
                with st.spinner("저장 중..."):
                    # Upload receipt if present
                    if uploaded_file is not None:
                        try:
                            file_bytes = uploaded_file.getvalue()
                            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            fname = f"receipt_{ts}_{uploaded_file.name}"
                            receipt_url = upload_to_drive(file_bytes, fname, uploaded_file.type)
                        except Exception as e:
                            st.warning(f"영수증 업로드 실패: {e}")
                            receipt_url = ""

                    # Append to Google Sheet
                    ws = get_sheet()
                    ws.append_row([
                        str(date),
                        category,
                        description,
                        amount,
                        payment_method,
                        receipt_url,
                        ocr_amount,
                        submitted_by,
                        datetime.datetime.now().isoformat(),
                    ])

                    st.cache_data.clear()
                    st.success("✅ 거래가 성공적으로 저장되었습니다!")
                    st.balloons()


# ──────────────────────────────────────────────
# PAGE: Transaction History
# ──────────────────────────────────────────────
elif page == "📋 거래 내역":
    st.markdown('<p class="main-header">거래 내역</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">전체 지출 기록을 조회합니다</p>', unsafe_allow_html=True)

    df = load_transactions()

    if df.empty:
        st.info("등록된 거래가 없습니다.")
    else:
        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            cats = ["전체"] + sorted(df["Category"].unique().tolist())
            sel_cat = st.selectbox("카테고리 필터", cats)
        with col_f2:
            if "Date" in df.columns and df["Date"].notna().any():
                min_date = df["Date"].min().date()
                max_date = df["Date"].max().date()
                date_range = st.date_input("기간", value=(min_date, max_date))
            else:
                date_range = None
        with col_f3:
            search = st.text_input("🔍 검색", placeholder="설명 검색...")

        filtered = df.copy()
        if sel_cat != "전체":
            filtered = filtered[filtered["Category"] == sel_cat]
        if date_range and len(date_range) == 2 and "Date" in filtered.columns:
            filtered = filtered[
                (filtered["Date"].dt.date >= date_range[0]) &
                (filtered["Date"].dt.date <= date_range[1])
            ]
        if search:
            filtered = filtered[
                filtered["Description"].astype(str).str.contains(search, case=False, na=False)
            ]

        st.markdown(f"**{len(filtered)}건** 조회됨 · 총 금액: **₩{filtered['Amount'].sum():,.0f}**")

        # Display table
        display_df = filtered.copy()
        if "Date" in display_df.columns:
            display_df["Date"] = display_df["Date"].dt.strftime("%Y-%m-%d")
        if "Amount" in display_df.columns:
            display_df["Amount"] = display_df["Amount"].apply(lambda x: f"₩{x:,.0f}")

        st.dataframe(
            display_df[["Date", "Category", "Description", "Amount", "Payment Method", "Submitted By"]],
            width="stretch",
            hide_index=True,
        )


# ──────────────────────────────────────────────
# PAGE: Budget Settings
# ──────────────────────────────────────────────
elif page == "⚙️ 예산 설정":
    st.markdown('<p class="main-header">예산 설정</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">카테고리별 월 예산을 설정합니다</p>', unsafe_allow_html=True)

    budgets = load_budgets()

    if not budgets.empty:
        st.markdown('<p class="section-title">현재 예산</p>', unsafe_allow_html=True)
        display_b = budgets.copy()
        display_b["Monthly Budget"] = display_b["Monthly Budget"].apply(lambda x: f"₩{x:,.0f}")
        st.dataframe(display_b, width="stretch", hide_index=True)

        # ── Edit existing budget ──
        st.markdown('<p class="section-title">예산 수정</p>', unsafe_allow_html=True)
        edit_idx = st.selectbox(
            "수정할 항목 선택",
            range(len(budgets)),
            format_func=lambda i: budgets.iloc[i]["Category"],
            key="edit_select",
        )
        sel = budgets.iloc[edit_idx]
        with st.form("edit_budget_form"):
            ec1, ec2, ec3 = st.columns(3)
            with ec1:
                edit_cat = st.text_input("카테고리", value=sel["Category"])
            with ec2:
                edit_budget = st.number_input("월 예산 (원)", value=int(sel["Monthly Budget"]), min_value=0, step=10000)
            with ec3:
                edit_year = st.text_input("연도", value=str(sel.get("Year", str(datetime.date.today().year))))
            edit_notes = st.text_input("메모", value=str(sel.get("Notes", "")))

            fc1, fc2 = st.columns(2)
            with fc1:
                if st.form_submit_button("💾 수정 저장", width="stretch"):
                    ws = get_budget_sheet()
                    sheet_row = edit_idx + 2
                    ws.update(f"A{sheet_row}:D{sheet_row}", [[edit_cat, edit_budget, edit_year, edit_notes]])
                    st.cache_data.clear()
                    st.success(f"✅ '{edit_cat}' 예산이 수정되었습니다.")
                    st.rerun()
            with fc2:
                if st.form_submit_button("🗑️ 삭제", width="stretch"):
                    ws = get_budget_sheet()
                    sheet_row = edit_idx + 2
                    ws.delete_rows(sheet_row)
                    st.cache_data.clear()
                    st.success(f"✅ '{sel['Category']}' 예산이 삭제되었습니다.")
                    st.rerun()

    st.markdown('<p class="section-title">예산 추가</p>', unsafe_allow_html=True)

    with st.form("budget_form"):
        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            new_cat = st.text_input("카테고리명", placeholder="예: 악기/장비")
        with bc2:
            new_budget = st.number_input("월 예산 (원)", min_value=0, step=10000)
        with bc3:
            new_year = st.text_input("연도", value=str(datetime.date.today().year))

        notes = st.text_input("메모", placeholder="선택사항")

        if st.form_submit_button("➕ 예산 추가", width="stretch"):
            if new_cat and new_budget > 0:
                ws = get_budget_sheet()
                ws.append_row([new_cat, new_budget, new_year, notes])
                st.cache_data.clear()
                st.success(f"✅ '{new_cat}' 예산이 추가되었습니다.")
                st.rerun()
            else:
                st.error("카테고리명과 예산 금액을 입력해주세요.")


# ──────────────────────────────────────────────
# PAGE: Report Download
# ──────────────────────────────────────────────
elif page == "📥 리포트 다운로드":
    st.markdown('<p class="main-header">리포트 다운로드</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Excel 형식의 상세 리포트를 다운로드합니다</p>', unsafe_allow_html=True)

    df = load_transactions()

    if df.empty:
        st.info("다운로드할 거래 데이터가 없습니다.")
    else:
        st.markdown(f"총 **{len(df)}건**의 거래 데이터가 포함됩니다.")

        # Summary preview
        if "Category" in df.columns:
            st.markdown('<p class="section-title">카테고리별 요약</p>', unsafe_allow_html=True)
            summary = df.groupby("Category")["Amount"].agg(["sum", "count"]).reset_index()
            summary.columns = ["카테고리", "총액", "건수"]
            summary["총액"] = summary["총액"].apply(lambda x: f"₩{x:,.0f}")
            st.dataframe(summary, width="stretch", hide_index=True)

        excel_data = generate_excel_report(df)
        today = datetime.date.today().strftime("%Y%m%d")

        st.download_button(
            label="📥 Excel 리포트 다운로드",
            data=excel_data,
            file_name=f"고등부_예산리포트_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.officedocument",
            width="stretch",
        )
